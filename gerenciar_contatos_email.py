"""Gerencia contatos de email armazenados em uma planilha XLSX externa."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Tuple

from config import PROJECT_ROOT

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class GerenciadorContatosEmail:
    CONFIG_KEY = "email_contatos"
    REQUIRED_COLUMNS = [
        "NOME_REGIONAL",
        "NOME_REG_FORTI",
        "NOME_DIRETOR_1",
        "NOME_GERENTE",
        "NOME_APOIO_1",
        "NOME_APOIO_2",
        "EMAIL_DIRETOR",
        "EMAIL_GERENTE",
        "EMAIL_APOIO_1",
        "EMAIL_APOIO_2",
    ]

    def __init__(self, environment_file: Path | None = None):
        self.environment_file = Path(environment_file or (PROJECT_ROOT / "environment.json"))

    def _load_environment(self) -> Dict:
        if not self.environment_file.exists():
            return {}
        with open(self.environment_file, "r", encoding="utf-8") as file:
            return json.load(file)

    def _save_environment(self, config: Dict) -> None:
        with open(self.environment_file, "w", encoding="utf-8") as file:
            json.dump(config, file, indent=2, ensure_ascii=False)

    def obter_configuracao(self) -> Dict[str, str]:
        config = self._load_environment().get(self.CONFIG_KEY, {})
        return {
            "xlsx_path": str(config.get("xlsx_path", "") or ""),
            "sheet_name": str(config.get("sheet_name", "") or ""),
        }

    def salvar_configuracao(self, xlsx_path: str, sheet_name: str = "") -> Dict[str, str]:
        environment = self._load_environment()
        environment[self.CONFIG_KEY] = {
            "xlsx_path": str((xlsx_path or "").strip()),
            "sheet_name": str((sheet_name or "").strip()),
        }
        self._save_environment(environment)
        return self.obter_configuracao()

    def _ensure_dependencies(self) -> None:
        if pd is None:
            raise RuntimeError("Pandas não está disponível para leitura da planilha de contatos.")
        if load_workbook is None:
            raise RuntimeError("openpyxl não está disponível para atualização da planilha de contatos.")

    def _resolve_planilha(self) -> Tuple[Path, str]:
        self._ensure_dependencies()
        config = self.obter_configuracao()
        xlsx_path = Path(config.get("xlsx_path", "")).expanduser()
        if not str(xlsx_path).strip():
            raise ValueError("Caminho da planilha de contatos não configurado.")
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Planilha não encontrada: {xlsx_path}")
        return xlsx_path, config.get("sheet_name", "")

    def _resolve_sheet_title(self, workbook, sheet_name: str) -> str:
        if sheet_name and sheet_name in workbook.sheetnames:
            return sheet_name
        return workbook.sheetnames[0]

    def listar_registros(self) -> List[Dict[str, str]]:
        xlsx_path, sheet_name = self._resolve_planilha()
        dataframe = pd.read_excel(xlsx_path, sheet_name=sheet_name or 0, dtype=str).fillna("")

        missing_columns = [column for column in self.REQUIRED_COLUMNS if column not in dataframe.columns]
        if missing_columns:
            raise ValueError(
                "A planilha não contém as colunas obrigatórias: " + ", ".join(missing_columns)
            )

        registros = []
        for row_index, row in dataframe.iterrows():
            registro = {column: str(row.get(column, "") or "").strip() for column in self.REQUIRED_COLUMNS}
            registro["_row_index"] = int(row_index)
            registros.append(registro)
        return registros

    def atualizar_registro(self, row_index: int, dados: Dict[str, str]) -> Dict[str, str]:
        xlsx_path, configured_sheet_name = self._resolve_planilha()
        workbook = load_workbook(xlsx_path)
        sheet_title = self._resolve_sheet_title(workbook, configured_sheet_name)
        worksheet = workbook[sheet_title]

        header_map = {}
        for column_index in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=column_index).value
            header_text = str(header_value or "").strip()
            if header_text:
                header_map[header_text] = column_index

        missing_columns = [column for column in self.REQUIRED_COLUMNS if column not in header_map]
        if missing_columns:
            raise ValueError(
                "A planilha não contém as colunas obrigatórias: " + ", ".join(missing_columns)
            )

        excel_row = int(row_index) + 2
        if excel_row > worksheet.max_row + 1:
            raise IndexError("Linha da planilha inválida para atualização.")

        registro_atualizado = {"_row_index": int(row_index)}
        for column in self.REQUIRED_COLUMNS:
            value = str(dados.get(column, "") or "").strip()
            worksheet.cell(row=excel_row, column=header_map[column]).value = value
            registro_atualizado[column] = value

        workbook.save(xlsx_path)
        workbook.close()
        return registro_atualizado
