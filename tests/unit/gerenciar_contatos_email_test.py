import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from gerenciar_contatos_email import GerenciadorContatosEmail


class GerenciadorContatosEmailTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.manager = GerenciadorContatosEmail(Path(self.temp_dir.name) / "environment.json")
        self.registro = {
            "NOME_REGIONAL": "PADRAO SLA",
            "NOME_REG_FORTI": "REPORT DE SEGURANCA-FGT_ORMEC_PARA",
            "NOME_DIRETOR_1": "Maria Silva",
            "EMAIL_DIRETOR": "maria@example.com",
            "NOME_GERENTE": "SEM_GERENTE",
            "EMAIL_GERENTE": "SEM_GERENTE",
            "NOME_APOIO_1": "Joao Souza",
            "EMAIL_APOIO_1": "joao@example.com",
            "NOME_APOIO_2": "",
            "EMAIL_APOIO_2": "",
        }

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_localiza_pela_correspondencia_forti(self):
        with patch.object(self.manager, "listar_registros", return_value=[self.registro]):
            result = self.manager.localizar_registro(nome_reg_forti="report_de seguranca-fgt ormec para")
        self.assertEqual(result["NOME_REGIONAL"], "PADRAO SLA")

    def test_monta_destinatarios_com_primeiros_nomes(self):
        result = self.manager.montar_destinatarios(self.registro)
        self.assertEqual(result["emails"], ["maria@example.com", "joao@example.com"])
        self.assertEqual([item["primeiro_nome"] for item in result["contatos"]], ["Maria", "Joao"])

    def test_cadastra_regional_em_nova_linha(self):
        xlsx_path = Path(self.temp_dir.name) / "contatos.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(self.manager.REQUIRED_COLUMNS)
        workbook.save(xlsx_path)
        workbook.close()
        self.manager.salvar_configuracao(str(xlsx_path))

        result = self.manager.cadastrar_registro(self.registro)

        workbook = load_workbook(xlsx_path, read_only=True)
        saved = workbook.active
        self.assertEqual(result["_row_index"], 0)
        self.assertEqual(saved.cell(row=2, column=1).value, "PADRAO SLA")
        self.assertEqual(saved.cell(row=2, column=2).value, "REPORT DE SEGURANCA-FGT_ORMEC_PARA")
        workbook.close()

    def test_impede_regional_duplicada(self):
        with patch.object(self.manager, "listar_registros", return_value=[self.registro]):
            with self.assertRaisesRegex(ValueError, "já está cadastrada"):
                self.manager.cadastrar_registro(self.registro)


if __name__ == "__main__":
    unittest.main()
